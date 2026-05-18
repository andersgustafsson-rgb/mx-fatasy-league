"""
Tid/övertid: månadsflikar, klistra in på «Klistra_in», välj månad, kör skript.

Skriptet kopierar inklistrat till vald månadsflik, bygger «Diagram_timmar» och «Graf_timmar».

Usage:
  py -3 scripts/feb_overtime_bar_charts.py --mall
      → skapar excel/Tidrapport_mall_mottagare.xlsx (tom mall för utskick, ingen data krävs)
  py -3 scripts/feb_overtime_bar_charts.py
  py -3 scripts/feb_overtime_bar_charts.py "C:\\path\\arbetsbok.xlsx"
  py -3 scripts/feb_overtime_bar_charts.py "C:\\path\\arbetsbok.xlsx" "Mars 2026"

Om källfilen redan heter *_diagram.xlsx skrivs samma fil över.
"""
from __future__ import annotations

import calendar
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

PASTE_HEADERS: tuple[str, ...] = (
    "Namn",
    "Datum",
    "Kl Fom",
    "Kl Tom",
    "Kl rast",
    "Rast",
    "Typ",
    "Orsak",
    "Bemanningstyp",
    "Proc",
    "Organisation",
    "Kto",
    "Tst",
    "Bev",
    "Bvä",
    "Med",
)

MONTHS_SV: tuple[str, ...] = (
    "Januari",
    "Februari",
    "Mars",
    "April",
    "Maj",
    "Juni",
    "Juli",
    "Augusti",
    "September",
    "Oktober",
    "November",
    "December",
)

SHEET_PASTE = "Klistra_in"
SHEET_MONTH_LIST = "_Månader"
SHEET_DATA = "Månadsdata"
SHEET_LEGACY = "februari"
# På Windows krockar «februari» med månadsfliken «Februari» — gamla exporten byts namn till detta.
SHEET_LEGACY_ARCHIVE = "Export_februari"
SHEET_DIAGRAM = "Diagram_timmar"
SHEET_CHART = "Graf_timmar"
SHEET_HELP = "Instruktioner"

# Rad på Klistra_in där rubriker alltid finns (användaren kan klistra ovanpå rad 8–…)
KLISTRA_HEADER_ROW = 8


def clean_str(val: object | None) -> str:
    if val is None:
        return ""
    return unicodedata.normalize("NFKC", str(val)).replace("\xa0", " ").strip()


def parse_time(val: object | None) -> float | None:
    s = clean_str(val)
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    return h + mi / 60.0


def parse_rast_minutes(val: object | None) -> float:
    s = clean_str(val)
    if not s:
        return 0.0
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return 0.0


def duration_hours(fom: float | None, tom: float | None) -> float:
    if fom is None or tom is None:
        return 0.0
    delta = tom - fom
    if delta <= 0:
        delta += 24.0
    return delta


ORSAC_ORDER = [
    "Poäng inst kort ledig",
    "Poäng åter StBy/Förskj",
    "Arbtid kontering",
    "Stand by schema",
    "Stand by ledig",
]


def normalize_orsak(raw: str) -> str:
    r = clean_str(raw)
    if not r:
        return "Övrigt"
    low = r.lower()
    if "poäng inst" in low or "pong inst" in low or "kort ledig" in low:
        return ORSAC_ORDER[0]
    if "stby" in low or "förskj" in low or "forskj" in low or "ater st" in low:
        return ORSAC_ORDER[1]
    if "kontering" in low or "arbtid" in low:
        return ORSAC_ORDER[2]
    if "stand by schema" in low or ("stand" in low and "schema" in low):
        return ORSAC_ORDER[3]
    if "stand by ledig" in low or ("stand" in low and "ledig" in low):
        return ORSAC_ORDER[4]
    return r or "Övrigt"


def find_header_row_any(ws: Worksheet, max_scan: int = 500) -> int | None:
    for r in range(1, min(ws.max_row, max_scan) + 1):
        if clean_str(ws.cell(r, 1).value) == "Namn":
            return r
    return None


def month_sheet_from_choice(val: object | None) -> str:
    s = clean_str(val)
    if not s:
        raise SystemExit(f"Välj månad i cell B1 på fliken «{SHEET_PASTE}» (listruta).")
    low = s.lower()
    for m in MONTHS_SV:
        if m.lower() == low:
            return m
    raise SystemExit(
        f"Ogiltig månad «{s}» i B1 på «{SHEET_PASTE}». Använd listan (t.ex. Februari, Mars)."
    )


def month_sheet_has_body(ws: Worksheet, hdr_row: int = 1) -> bool:
    for r in range(hdr_row + 1, min(ws.max_row, 8000) + 1):
        if clean_str(ws.cell(r, 1).value):
            return True
    return False


def clear_sheet_below_header(ws: Worksheet, hdr_row: int = 1, ncols: int = len(PASTE_HEADERS)) -> None:
    for r in range(hdr_row + 1, ws.max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(r, c).value = None


def write_paste_headers_row(ws: Worksheet, row: int) -> None:
    for i, h in enumerate(PASTE_HEADERS, start=1):
        c = ws.cell(row, i, h)
        c.font = Font(bold=True)


def archive_legacy_februari_sheet(out_wb: Workbook) -> None:
    """Byt namn på gammal «februari» så att månadsfliken «Februari» kan skapas (Windows är skiftlägesokänslig)."""
    if SHEET_LEGACY not in out_wb.sheetnames:
        return
    out_wb[SHEET_LEGACY].title = SHEET_LEGACY_ARCHIVE


def fix_februari_tab_name(out_wb: Workbook) -> None:
    """Om «Februari» saknas men «Februari1» finns (tidigare krock), byt tillbaka namnet."""
    if "Februari" in out_wb.sheetnames:
        return
    for cand in ("Februari1", "Februari2"):
        if cand in out_wb.sheetnames:
            out_wb[cand].title = "Februari"
            return


def remove_empty_februari_dup_tabs(out_wb: Workbook) -> None:
    """Ta bort tomma «Februari2»-flikar efter tidigare namnkrock."""
    for name in list(out_wb.sheetnames):
        if not name.startswith("Februari") or name == "Februari":
            continue
        ws = out_wb[name]
        if month_sheet_has_body(ws, 1):
            continue
        del out_wb[name]


def ensure_month_tabs(out_wb: Workbook) -> None:
    archive_legacy_februari_sheet(out_wb)
    fix_februari_tab_name(out_wb)
    remove_empty_februari_dup_tabs(out_wb)
    for m in MONTHS_SV:
        if m not in out_wb.sheetnames:
            ws = out_wb.create_sheet(m)
            write_paste_headers_row(ws, 1)


def ensure_month_list_sheet(out_wb: Workbook) -> None:
    if SHEET_MONTH_LIST in out_wb.sheetnames:
        ws = out_wb[SHEET_MONTH_LIST]
    else:
        ws = out_wb.create_sheet(SHEET_MONTH_LIST)
    for i, m in enumerate(MONTHS_SV, start=1):
        ws.cell(i, 1, m)
    ws.sheet_state = "hidden"


def ensure_klistra_in(out_wb: Workbook) -> Worksheet:
    if SHEET_PASTE in out_wb.sheetnames:
        return out_wb[SHEET_PASTE]
    ws = out_wb.create_sheet(SHEET_PASTE)
    ws["A1"] = "1) Välj månad här"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Februari"
    ws["A3"] = (
        "2) Klistra in exporten: börja på rad 8 så att «Namn» står i A8 "
        "(du kan ersätta rubrikraden; datarader följer under)."
    )
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 18
    write_paste_headers_row(ws, KLISTRA_HEADER_ROW)

    ensure_month_list_sheet(out_wb)
    dv = DataValidation(
        type="list",
        formula1=f"={SHEET_MONTH_LIST}!$A$1:$A${len(MONTHS_SV)}",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Ogiltig månad",
        error="Välj en månad från listan.",
    )
    ws.add_data_validation(dv)
    dv.add(ws["B1"])
    return ws


def migrate_legacy_february(out_wb: Workbook) -> None:
    """Om arkiverad export finns och «Februari» är tom — kopiera dit."""
    if "Februari" not in out_wb.sheetnames or SHEET_LEGACY_ARCHIVE not in out_wb.sheetnames:
        return
    feb_new = out_wb["Februari"]
    if month_sheet_has_body(feb_new, 1):
        return
    leg = out_wb[SHEET_LEGACY_ARCHIVE]
    hdr = find_header_row_any(leg)
    if hdr is None:
        return
    write_paste_headers_row(feb_new, 1)
    r_dst = 2
    for r_src in range(hdr + 1, leg.max_row + 1):
        if not any(leg.cell(r_src, c).value for c in range(1, len(PASTE_HEADERS) + 1)):
            continue
        for c in range(1, len(PASTE_HEADERS) + 1):
            feb_new.cell(r_dst, c).value = leg.cell(r_src, c).value
        r_dst += 1


def migrate_manadsdata_to_february(out_wb: Workbook) -> None:
    if SHEET_DATA not in out_wb.sheetnames or "Februari" not in out_wb.sheetnames:
        return
    md, feb = out_wb[SHEET_DATA], out_wb["Februari"]
    if clean_str(md.cell(1, 1).value) != "Namn" or not month_sheet_has_body(md, 1):
        return
    if month_sheet_has_body(feb, 1):
        return
    for r in range(1, md.max_row + 1):
        for c in range(1, len(PASTE_HEADERS) + 1):
            feb.cell(r, c).value = md.cell(r, c).value


def find_klistra_header_row(ws: Worksheet) -> int:
    """Rubrik «Namn» på rad 8 eller var användaren klistrat (sök uppåt/nedåt från rad 8)."""
    fixed = KLISTRA_HEADER_ROW
    if clean_str(ws.cell(fixed, 1).value) == "Namn":
        return fixed
    for r in range(5, min(ws.max_row, 80) + 1):
        if clean_str(ws.cell(r, 1).value) == "Namn":
            return r
    return fixed


def flush_klistra_to_month_sheet(out_wb: Workbook) -> bool:
    """
    Om Klistra_in har datarader under rubriken: kopiera rubrik+data till vald månadsflik
    och töm inklistrat (lämnar B1 och rubrikrad 8).
    """
    if SHEET_PASTE not in out_wb.sheetnames:
        return False
    kl = out_wb[SHEET_PASTE]
    month_name = month_sheet_from_choice(kl["B1"].value)
    hdr = find_klistra_header_row(kl)
    if not any(clean_str(kl.cell(r, 1).value) for r in range(hdr + 1, min(kl.max_row, hdr + 8000) + 1)):
        return False
    tgt = out_wb[month_name]
    clear_sheet_below_header(tgt, 1)
    write_paste_headers_row(tgt, 1)
    r_dst = 1
    for r_src in range(hdr, kl.max_row + 1):
        if r_src > hdr and not any(kl.cell(r_src, c).value for c in range(1, len(PASTE_HEADERS) + 1)):
            continue
        if not any(kl.cell(r_src, c).value for c in range(1, len(PASTE_HEADERS) + 1)):
            continue
        for c in range(1, len(PASTE_HEADERS) + 1):
            tgt.cell(r_dst, c).value = kl.cell(r_src, c).value
        r_dst += 1
    # Töm klistret men behåll layout
    for r in range(hdr + 1, kl.max_row + 1):
        for c in range(1, len(PASTE_HEADERS) + 1):
            kl.cell(r, c).value = None
    write_paste_headers_row(kl, KLISTRA_HEADER_ROW)
    return True


def aggregate_totals(ws: Worksheet, hdr_row: int) -> dict[str, dict[str, float]]:
    totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = clean_str(ws.cell(r, 1).value)
        if not name:
            continue
        fom = parse_time(ws.cell(r, 3).value)
        tom = parse_time(ws.cell(r, 4).value)
        rast_min = parse_rast_minutes(ws.cell(r, 6).value)
        gross = duration_hours(fom, tom)
        net = max(0.0, gross - rast_min / 60.0)
        orsak = normalize_orsak(clean_str(ws.cell(r, 8).value))
        totals[name][orsak] += net
    return totals


def parse_month_label(ws: Worksheet, hdr_row: int) -> str | None:
    for r in range(hdr_row + 1, min(ws.max_row, hdr_row + 2000) + 1):
        raw = ws.cell(r, 2).value
        if raw is None:
            continue
        if isinstance(raw, datetime):
            d = raw.date()
        else:
            s = clean_str(raw)
            if not s:
                continue
            try:
                d = datetime.strptime(s[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
        swe = {
            1: "januari",
            2: "februari",
            3: "mars",
            4: "april",
            5: "maj",
            6: "juni",
            7: "juli",
            8: "augusti",
            9: "september",
            10: "oktober",
            11: "november",
            12: "december",
        }
        return f"{swe.get(d.month, calendar.month_name[d.month]).capitalize()} {d.year}"
    return None


def resolve_data_legacy(wb) -> tuple[Worksheet, int]:
    if SHEET_DATA in wb.sheetnames:
        md = wb[SHEET_DATA]
        if clean_str(md.cell(1, 1).value) != "Namn":
            raise SystemExit(f"«{SHEET_DATA}»: cell A1 ska vara «Namn».")
        if month_sheet_has_body(md, 1):
            return md, 1
        raise SystemExit(
            f"«{SHEET_DATA}» saknar data under rad 1. Använd «{SHEET_PASTE}» eller ta bort fliken för att falla tillbaka."
        )
    if SHEET_LEGACY_ARCHIVE in wb.sheetnames:
        feb = wb[SHEET_LEGACY_ARCHIVE]
        hr = find_header_row_any(feb)
        if hr is not None:
            return feb, hr
    for name in wb.sheetnames:
        if name in (
            SHEET_HELP,
            SHEET_DIAGRAM,
            SHEET_CHART,
            SHEET_MONTH_LIST,
            SHEET_PASTE,
            SHEET_LEGACY_ARCHIVE,
        ):
            continue
        if name in MONTHS_SV:
            continue
        ws = wb[name]
        hr = find_header_row_any(ws)
        if hr is not None and hr != 1:
            if any(clean_str(ws.cell(r, 1).value) for r in range(hr + 1, min(ws.max_row, hr + 500) + 1)):
                return ws, hr
    raise SystemExit("Hittade ingen data. Använd «Klistra_in» eller lägg rader under «Namn» på en månadsflik.")


def compute_out_path(src: Path) -> Path:
    if src.stem.endswith("_diagram"):
        return src
    return src.with_name(src.stem + "_diagram.xlsx")


def find_input_path(argv: list[str]) -> Path:
    if len(argv) > 1:
        return Path(argv[1]).expanduser().resolve()
    downloads = Path.home() / "Downloads"
    diagram_hits = sorted(downloads.glob("*feb*sammanfattning*2026*_diagram.xlsx"))
    if diagram_hits:
        return diagram_hits[-1]
    hits = sorted(downloads.glob("*feb*sammanfattning*2026*.xlsx"))
    if hits:
        return hits[-1]
    raise SystemExit("Hittade ingen xlsx i Downloads. Ange sökväg som argument.")


def remove_output_artifacts(out_wb: Workbook) -> None:
    for name in (SHEET_DIAGRAM, SHEET_CHART, "Diagram_februari", "Graf_februari"):
        if name in out_wb.sheetnames:
            del out_wb[name]


def copy_legacy_table_to_manadsdata(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    hdr = find_header_row_any(src_ws)
    if hdr is None:
        return
    write_paste_headers_row(dst_ws, 1)
    r_dst = 2
    for r_src in range(hdr + 1, src_ws.max_row + 1):
        if not any(src_ws.cell(r_src, c).value for c in range(1, len(PASTE_HEADERS) + 1)):
            continue
        for c in range(1, len(PASTE_HEADERS) + 1):
            dst_ws.cell(r_dst, c).value = src_ws.cell(r_src, c).value
        r_dst += 1


def ensure_manadsdata_sheet(out_wb: Workbook) -> None:
    """Behålls för bakåtkompabilitet; skapas bara om den saknas och gammal källa finns."""
    if SHEET_DATA in out_wb.sheetnames:
        return
    if SHEET_LEGACY_ARCHIVE not in out_wb.sheetnames:
        return
    md = out_wb.create_sheet(SHEET_DATA, len(out_wb.sheetnames))
    write_paste_headers_row(md, 1)
    copy_legacy_table_to_manadsdata(out_wb[SHEET_LEGACY_ARCHIVE], md)


def build_instructions_text(month_label: str) -> str:
    h = "\n".join(PASTE_HEADERS)
    months = ", ".join(MONTHS_SV)
    return f"""VAD DEN HÄR ARBETSBOKEN GÖR
• Tolv färdiga flikar för året: {months}
• Fliken «{SHEET_PASTE}»: välj månad i B1 och klistra in exporten (rubrik «Namn» ska ligga i kolumn A, börja på rad {KLISTRA_HEADER_ROW}).
• «{SHEET_DIAGRAM}» = tabell per person/status, «{SHEET_CHART}» = diagram.
• Excel utan Python: läs excel/LAS_MIG_FORST.txt (samma mapp som ModulTidrapport.bas). Spara som .xlsm, lägg in makro, knapp «Flytta till vald månad».

AKTUELL RUBRIK I DIAGRAM (uppskattad)
• {month_label}

SÅ HÄR ANVÄNDER DU «{SHEET_PASTE}» (i Excel)
1) Välj månad i B1 (listruta).
2) Klistra in så att rubrikraden (med «Namn» i A-kolumnen) ligger på rad {KLISTRA_HEADER_ROW}.
3) Datat följer på raderna under.
4) Kör makrot «FlyttaTillValdManad» (eller motsvarande knapp). Då kopieras blocket till vald månadsflik. Väljer du bara månad utan att köra makro händer inget med flytten.

KOLUMNER (samma som systemexport)
{h}

ÄLDRE EXPORT / «{SHEET_DATA}»
• Gammal månadsflik byts namn till «{SHEET_LEGACY_ARCHIVE}» (så «Februari» kan användas som månad). Data kan migreras till «Februari».
• «{SHEET_DATA}» skapas vid behov som kopia av arkivfliken.

VALFRITT (Python i utvecklingsmiljö)
  py -3 scripts\\feb_overtime_bar_charts.py \"<sökväg till denna xlsx>\"
"""


def write_instructions_sheet(ws: Worksheet, month_label: str) -> None:
    ws["A1"] = build_instructions_text(month_label)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 108
    ws.row_dimensions[1].height = 520


MOTTAGARE_INSTRUKTION = """\
SÅ HÄR ANVÄNDER DU ARBETSBOKEN

1) Öppna filen i Microsoft Excel. Om du frågas om makron — tillåt för den här filen (annars kan inte flytt/diagram köras).

2) Gå till fliken «Klistra_in».

3) Välj månad i cell B1 (Januari … December).

4) Klistra in exporten så att rubriken «Namn» står i kolumn A (gärna rad 8; datarader rakt under).

5) Tryck Alt+F8. Markera makrot «FlyttaTillValdManad». Klicka Kör.

Då kopieras datan till rätt månadsflik. Tabellen «Diagram_timmar» och diagrammet «Graf_timmar» skapas eller uppdateras.

────────────────────────────────────
DU SOM SKICKAR MALLEN TILL ANDRA (en gång)
Spara som «Excel-arbetsbok med makro (*.xlsm)». Lägg in makrokoden från excel/ModulTidrapport.bas enligt excel/LAS_MIG_FORST.txt.
Skicka sedan .xlsm till mottagarna — de behöver inte Python.
"""


def create_recipient_template_xlsx(out_path: Path) -> None:
    """Tom .xlsx: alla månadsflikar + Klistra_in + instruktion. Makro läggs in manuellt → .xlsm."""
    out_path = out_path.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    help_ws = wb.active
    help_ws.title = SHEET_HELP
    help_ws["A1"] = MOTTAGARE_INSTRUKTION
    help_ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    help_ws.column_dimensions["A"].width = 96
    help_ws.row_dimensions[1].height = 420

    ensure_month_tabs(wb)
    ensure_klistra_in(wb)
    reorder_sheets_front(wb)

    wb.save(out_path)


def reorder_sheets_front(out_wb: Workbook) -> None:
    """Tab-ordning: Instruktioner, Klistra_in, januari–december, sedan övrigt."""
    want_first = [SHEET_HELP, SHEET_PASTE, *MONTHS_SV]
    pos = 0
    for name in want_first:
        if name not in out_wb.sheetnames:
            continue
        sh = out_wb[name]
        idx = out_wb.sheetnames.index(name)
        delta = pos - idx
        if delta != 0:
            out_wb.move_sheet(sh, offset=delta)
        pos += 1


def build_stacked_bar_chart(
    dws: Worksheet,
    month_label: str,
    ordered_cols: list[str],
    names_sorted: list[str],
    header_row: int,
    sum_col: int,
) -> BarChart:
    first_data_row = header_row + 1
    last_data_row = header_row + len(names_sorted)
    chart = BarChart(barDir="bar", grouping="stacked", gapWidth=42, overlap=100)
    chart.title = f"{month_label} — timmar per person (per status)"
    chart.x_axis.title = "Timmar"
    chart.y_axis.title = None
    chart.legend.position = "r"
    chart.legend.overlay = False
    n = len(names_sorted)
    chart.height = min(52, 10 + max(n, 8) * 0.42)
    chart.width = 34
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.scaling.min = 0
    chart.x_axis.numFmt = "0.0"
    data_ref = Reference(
        dws,
        min_col=2,
        max_col=sum_col - 1,
        min_row=header_row,
        max_row=last_data_row,
    )
    cats_ref = Reference(dws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    # Textkategorier (namn) måste använda strRef — annars visas inte etiketter i Excel.
    cat_f = str(cats_ref)
    for s in chart.series:
        s.cat = AxDataSource(strRef=StrRef(f=cat_f))
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    return chart


def main() -> None:
    src = find_input_path(sys.argv)
    if not src.is_file():
        raise SystemExit(f"Hittar inte filen: {src}")

    month_override = clean_str(sys.argv[2]) if len(sys.argv) > 2 else ""

    out_path = compute_out_path(src)
    out_wb = load_workbook(src, data_only=False)

    ensure_month_tabs(out_wb)
    ensure_klistra_in(out_wb)
    migrate_legacy_february(out_wb)
    migrate_manadsdata_to_february(out_wb)
    ensure_manadsdata_sheet(out_wb)
    flush_klistra_to_month_sheet(out_wb)

    out_wb.save(out_path)

    wb_data = load_workbook(out_path, data_only=True)
    chart_ws_name = None
    if SHEET_PASTE in wb_data.sheetnames:
        chart_ws_name = month_sheet_from_choice(wb_data[SHEET_PASTE]["B1"].value)
    if chart_ws_name and chart_ws_name in wb_data.sheetnames:
        data_ws = wb_data[chart_ws_name]
        hdr_row = 1
        if not month_sheet_has_body(data_ws, 1):
            raise SystemExit(f"Ingen data på «{chart_ws_name}». Klistra in på «{SHEET_PASTE}» och kör igen.")
        month_label = month_override or parse_month_label(data_ws, 1) or chart_ws_name
    else:
        data_ws, hdr_row = resolve_data_legacy(wb_data)
        month_label = month_override or parse_month_label(data_ws, hdr_row) or "Aktuell månad"

    totals = aggregate_totals(data_ws, hdr_row)
    if not totals:
        raise SystemExit("Ingen data under rubrikraden.")

    extra_keys: set[str] = set()
    for m in totals.values():
        for k in m:
            if k not in ORSAC_ORDER:
                extra_keys.add(k)
    ordered_cols = list(ORSAC_ORDER) + sorted(k for k in extra_keys if k not in ORSAC_ORDER)

    out_wb = load_workbook(out_path, data_only=False)
    remove_output_artifacts(out_wb)

    if SHEET_HELP in out_wb.sheetnames:
        del out_wb[SHEET_HELP]
    help_ws = out_wb.create_sheet(SHEET_HELP, 0)
    write_instructions_sheet(help_ws, month_label)

    ensure_month_tabs(out_wb)
    ensure_klistra_in(out_wb)
    ensure_month_list_sheet(out_wb)
    reorder_sheets_front(out_wb)

    dws = out_wb.create_sheet(SHEET_DIAGRAM, len(out_wb.sheetnames))
    dws["A1"] = (
        f"{month_label} — timmar per person "
        "(Tom − Från, nattpass +24 h, rast «Rast» i minuter)"
    )
    dws["A1"].font = Font(bold=True, size=12)
    dws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(ordered_cols))

    header_row = 3
    dws.cell(header_row, 1, "Namn")
    for i, title in enumerate(ordered_cols, start=2):
        dws.cell(header_row, i, title)
    sum_col = 2 + len(ordered_cols)
    dws.cell(header_row, sum_col, "Summa h")

    names_sorted = sorted(totals.keys(), key=lambda n: (-sum(totals[n].values()), n))
    for idx, name in enumerate(names_sorted, start=1):
        row = header_row + idx
        dws.cell(row, 1, name)
        row_sum = 0.0
        for i, col_title in enumerate(ordered_cols, start=2):
            v = float(totals[name].get(col_title, 0.0))
            dws.cell(row, i, round(v, 2))
            row_sum += v
        dws.cell(row, sum_col, round(row_sum, 2))

    dws.column_dimensions["A"].width = 34
    for c in range(2, sum_col + 1):
        dws.column_dimensions[get_column_letter(c)].width = 16

    if names_sorted:
        chart = build_stacked_bar_chart(dws, month_label, ordered_cols, names_sorted, header_row, sum_col)
        cws = out_wb.create_chartsheet(SHEET_CHART)
        cws.add_chart(chart)

    out_wb.save(out_path)
    print(f"Sparad: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].strip().lower() in ("--mall", "-mall", "/mall"):
        mall_path = Path(__file__).resolve().parent.parent / "excel" / "Tidrapport_mall_mottagare.xlsx"
        create_recipient_template_xlsx(mall_path)
        print(f"Mall skapad: {mall_path}")
    else:
        main()
